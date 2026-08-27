import io
import logging
import xml.etree.ElementTree as ET
from decimal import Decimal, InvalidOperation
from typing import Dict, Any, List, Optional, Union
import openpyxl
from django.db import transaction

from colaboradores.models import Colaborador

logger = logging.getLogger(__name__)

def parse_decimal_safe(val: Any, default: Decimal = Decimal("0.00")) -> Decimal:
    """
    Converte valores numéricos/monetários com segurança para Decimal.
    """
    if val is None:
        return default
    texto = str(val).strip()
    if not texto or texto == "-" or texto == "  /  /    ":
        return default
    try:
        if "," in texto and "." in texto:
            texto = texto.replace(".", "").replace(",", ".")
        elif "," in texto:
            texto = texto.replace(",", ".")
        return Decimal(texto)
    except (ValueError, InvalidOperation, TypeError):
        return default


def extrair_dados_rescisao_xml(conteudo_bytes: bytes) -> List[Dict[str, Any]]:
    """
    Extrai dados da Página 2 do arquivo XML (SpreadsheetML) do relatório de rescisões/líquido do TOTVS.
    """
    root = ET.fromstring(conteudo_bytes)
    ns = {'ss': 'urn:schemas-microsoft-com:office:spreadsheet'}
    
    worksheets = root.findall('ss:Worksheet', ns)
    if not worksheets:
        worksheets = root.findall('.//Worksheet')
        
    if not worksheets:
        raise ValueError("Nenhuma planilha encontrada no arquivo XML.")
    
    # Seleciona a página 2 (índice 1) ou a planilha '2-Func Benef'
    ws_target = None
    if len(worksheets) >= 2:
        ws_target = worksheets[1]
    else:
        for ws in worksheets:
            name = ws.attrib.get('{urn:schemas-microsoft-com:office:spreadsheet}Name') or ws.attrib.get('Name') or ''
            if 'FUNC' in name.upper() or 'BENEF' in name.upper() or '2' in name:
                ws_target = ws
                break
        if not ws_target:
            ws_target = worksheets[0]
            
    rows = ws_target.findall('.//ss:Row', ns) or ws_target.findall('.//Row')
    if not rows:
        raise ValueError("Nenhuma linha encontrada na página 2 do relatório de rescisões.")
        
    # Localiza a linha de cabeçalho
    header_row_idx = None
    headers = {}
    
    for idx, row in enumerate(rows):
        cells_map = {}
        col = 1
        for cell in (row.findall('ss:Cell', ns) or row.findall('Cell')):
            idx_attr = cell.attrib.get('{urn:schemas-microsoft-com:office:spreadsheet}Index') or cell.attrib.get('Index')
            if idx_attr:
                col = int(idx_attr)
            data = cell.find('ss:Data', ns) if cell.find('ss:Data', ns) is not None else cell.find('Data')
            if data is not None and data.text:
                cells_map[col - 1] = data.text.strip()
            col += 1
            
        values_upper = [v.upper() for v in cells_map.values()]
        if any('MATRICULA' in v for v in values_upper) and (any('VALOR' in v for v in values_upper) or any('LIQUIDO' in v for v in values_upper)):
            header_row_idx = idx
            headers = cells_map
            break
            
    if header_row_idx is None:
        raise ValueError("Cabeçalho com 'Matricula' e 'Valor' não foi encontrado na página 2 do relatório.")
        
    mat_col = None
    val_col = None
    nome_col = None
    cpf_col = None
    cc_col = None
    filial_col = None
    
    for col_idx, col_name in headers.items():
        cn = col_name.upper()
        if 'MATRICULA' in cn:
            mat_col = col_idx
        elif 'VALOR' in cn or 'LIQUIDO' in cn:
            val_col = col_idx
        elif 'NOME' in cn:
            nome_col = col_idx
        elif 'CPF' in cn:
            cpf_col = col_idx
        elif 'CENTRO' in cn and 'CUSTO' in cn:
            cc_col = col_idx
        elif 'FILIAL' in cn:
            filial_col = col_idx
            
    if mat_col is None or val_col is None:
        raise ValueError("Não foi possível identificar as colunas de Matrícula ou Valor na página 2.")
        
    registros = []
    for r_idx in range(header_row_idx + 1, len(rows)):
        row = rows[r_idx]
        cells_map = {}
        col = 1
        for cell in (row.findall('ss:Cell', ns) or row.findall('Cell')):
            idx_attr = cell.attrib.get('{urn:schemas-microsoft-com:office:spreadsheet}Index') or cell.attrib.get('Index')
            if idx_attr:
                col = int(idx_attr)
            data = cell.find('ss:Data', ns) if cell.find('ss:Data', ns) is not None else cell.find('Data')
            if data is not None and data.text:
                cells_map[col - 1] = data.text.strip()
            col += 1
            
        mat = cells_map.get(mat_col, '').strip()
        val_str = cells_map.get(val_col, '').strip()
        nome = cells_map.get(nome_col, '').strip() if nome_col is not None else ''
        
        if not mat or 'TOTAL' in mat.upper() or 'TOTAL' in nome.upper() or 'TOTAL' in val_str.upper():
            continue
            
        valor = parse_decimal_safe(val_str)
        registros.append({
            'matricula': mat,
            'valor': valor,
            'nome': nome,
            'cpf': cells_map.get(cpf_col, '').strip() if cpf_col is not None else '',
            'centro_custo': cells_map.get(cc_col, '').strip() if cc_col is not None else '',
            'filial': cells_map.get(filial_col, '').strip() if filial_col is not None else ''
        })
        
    return registros


def extrair_dados_rescisao_excel(arquivo_bytes: bytes) -> List[Dict[str, Any]]:
    """
    Extrai dados da Página 2 de uma planilha Excel (.xlsx, .xlsm, .xls).
    """
    wb = openpyxl.load_workbook(io.BytesIO(arquivo_bytes), data_only=True)
    if len(wb.sheetnames) >= 2:
        ws = wb.worksheets[1]
    else:
        ws = wb.worksheets[0]
        
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError("Planilha vazia.")
        
    header_row_idx = None
    headers = []
    for idx, row in enumerate(rows):
        row_str = [str(c or '').strip() for c in row]
        row_upper = [c.upper() for c in row_str]
        if any('MATRICULA' in c for c in row_upper) and (any('VALOR' in c for c in row_upper) or any('LIQUIDO' in c for c in row_upper)):
            header_row_idx = idx
            headers = row_str
            break
            
    if header_row_idx is None:
        raise ValueError("Cabeçalho com 'Matricula' e 'Valor' não foi encontrado na página 2.")
        
    mat_col = None
    val_col = None
    nome_col = None
    cpf_col = None
    cc_col = None
    filial_col = None
    
    for col_idx, col_name in enumerate(headers):
        cn = col_name.upper()
        if 'MATRICULA' in cn:
            mat_col = col_idx
        elif 'VALOR' in cn or 'LIQUIDO' in cn:
            val_col = col_idx
        elif 'NOME' in cn:
            nome_col = col_idx
        elif 'CPF' in cn:
            cpf_col = col_idx
        elif 'CENTRO' in cn and 'CUSTO' in cn:
            cc_col = col_idx
        elif 'FILIAL' in cn:
            filial_col = col_idx
            
    if mat_col is None or val_col is None:
        raise ValueError("Não foi possível identificar as colunas de Matrícula ou Valor na página 2.")
        
    registros = []
    for row in rows[header_row_idx + 1:]:
        if not row:
            continue
        mat = str(row[mat_col] or '').strip() if mat_col is not None and mat_col < len(row) else ''
        val_raw = row[val_col] if val_col is not None and val_col < len(row) else 0
        nome = str(row[nome_col] or '').strip() if nome_col is not None and nome_col < len(row) else ''
        
        if not mat or 'TOTAL' in mat.upper() or 'TOTAL' in nome.upper():
            continue
            
        valor = parse_decimal_safe(val_raw)
        registros.append({
            'matricula': mat,
            'valor': valor,
            'nome': nome,
            'cpf': str(row[cpf_col] or '').strip() if cpf_col is not None and cpf_col < len(row) else '',
            'centro_custo': str(row[cc_col] or '').strip() if cc_col is not None and cc_col < len(row) else '',
            'filial': str(row[filial_col] or '').strip() if filial_col is not None and filial_col < len(row) else ''
        })
        
    return registros


def importar_valores_rescisao_de_arquivo(
    arquivo_bytes: bytes,
    nome_arquivo: str = "",
    progress_callback=None
) -> Dict[str, Any]:
    """
    Processa o upload do arquivo de valores de rescisão (TOTVS Líquido / GPER020 XML / Excel) de forma assíncrona.
    Atualiza exclusivamente o campo valor_rescisao_estimado dos colaboradores no banco de dados
    a partir da Página 2 do relatório.
    """
    if not arquivo_bytes:
        logger.info("Importação abortada: conteúdo do arquivo de rescisão vazio.")
        return {
            "total": 0,
            "atualizados": 0,
            "valor_total": 0.0,
            "nao_encontrados": [],
        }

    if progress_callback:
        progress_callback(10, "Lendo e interpretando arquivo de rescisão (Página 2)...")

    # Detecta se é XML SpreadsheetML ou planilha Excel convencional
    primeiros_bytes = arquivo_bytes[:500].strip()
    is_xml = b"<?xml" in primeiros_bytes or b"<Workbook" in primeiros_bytes or nome_arquivo.lower().endswith(".xml")

    try:
        if is_xml:
            registros = extrair_dados_rescisao_xml(arquivo_bytes)
        else:
            try:
                registros = extrair_dados_rescisao_excel(arquivo_bytes)
            except Exception:
                # Tenta como XML se falhar no openpyxl
                registros = extrair_dados_rescisao_xml(arquivo_bytes)
    except Exception as exc:
        logger.error(f"Erro ao extrair dados de valores de rescisão: {exc}")
        raise ValueError(f"Não foi possível processar o relatório de rescisão: {str(exc)}")

    if not registros:
        return {
            "total": 0,
            "atualizados": 0,
            "valor_total": 0.0,
            "nao_encontrados": [],
        }

    if progress_callback:
        progress_callback(30, "Carregando base de colaboradores para cruzamento...")

    colaboradores_dict = {c.re: c for c in Colaborador.objects.all()}
    
    para_atualizar = []
    nao_encontrados = []
    valor_total_importado = Decimal("0.00")
    total_linhas = len(registros)

    if progress_callback:
        progress_callback(50, "Atualizando valores de rescisão dos colaboradores...")

    for idx, reg in enumerate(registros, start=1):
        mat = reg["matricula"]
        val = reg["valor"]
        valor_total_importado += val

        colab = colaboradores_dict.get(mat)
        if colab:
            if colab.valor_rescisao_estimado != val:
                colab.valor_rescisao_estimado = val
                para_atualizar.append(colab)
        else:
            nao_encontrados.append({
                "matricula": mat,
                "nome": reg["nome"],
                "valor": float(val),
                "centro_custo": reg["centro_custo"],
            })

        if progress_callback and idx % 500 == 0:
            progresso = 50 + int((idx / max(1, total_linhas)) * 35)  # 50% a 85%
            progress_callback(progresso, f"Processando registros... ({idx}/{total_linhas})")

    if progress_callback:
        progress_callback(88, "Persistindo novos valores de rescisão no banco de dados...")

    if para_atualizar:
        with transaction.atomic():
            Colaborador.objects.bulk_update(
                para_atualizar,
                ["valor_rescisao_estimado"],
                batch_size=4000
            )

    if progress_callback:
        progress_callback(100, "Valores de rescisão importados com sucesso!")

    return {
        "total": len(registros),
        "atualizados": len(para_atualizar),
        "valor_total": float(valor_total_importado),
        "nao_encontrados": nao_encontrados,
    }

# Alias para compatibilidade
importar_gper020_rescisao = importar_valores_rescisao_de_arquivo
