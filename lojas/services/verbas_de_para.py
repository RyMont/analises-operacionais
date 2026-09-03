import os
import json
from decimal import Decimal
from typing import Dict, Any, Optional

# Caminho do arquivo JSON permanente de De-Para
FIXTURE_PATH = os.path.join(
    os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
    "fixtures",
    "de_para_verbas.json"
)
FIXTURE_EXAMPLE_PATH = os.path.join(
    os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
    "fixtures",
    "de_para_verbas.json.example"
)

# Caminho da planilha original de verbas (configurável via PLANILHA_VERBAS_PATH no .env)
def obter_caminho_planilha() -> str:
    """
    Obtém o caminho da planilha de verbas de forma resiliente,
    consultando as configurações do Django, python-decouple ou variáveis de ambiente.
    """
    path = ""
    try:
        from django.conf import settings
        if hasattr(settings, "PLANILHA_VERBAS_PATH"):
            path = str(settings.PLANILHA_VERBAS_PATH).strip()
    except Exception:
        pass

    if not path:
        try:
            from decouple import config
            path = config("PLANILHA_VERBAS_PATH", default="").strip()
        except Exception:
            pass

    if not path:
        try:
            from pathlib import Path
            from decouple import Config, RepositoryEnv
            base_dir = Path(__file__).resolve().parent.parent.parent
            env_file = base_dir / ".env"
            if env_file.exists():
                cfg = Config(RepositoryEnv(str(env_file)))
                path = cfg("PLANILHA_VERBAS_PATH", default="").strip()
        except Exception:
            pass

    if not path:
        path = os.getenv("PLANILHA_VERBAS_PATH", "").strip()

    # Remove aspas se o usuário colocou no .env (ex: "F:/...")
    if (path.startswith('"') and path.endswith('"')) or (path.startswith("'") and path.endswith("'")):
        path = path[1:-1].strip()

    return path


# Cache em memória
_MAPA_DE_PARA: Optional[Dict[str, Dict[str, str]]] = None


def carregar_mapa_verbas(forcar_recarga: bool = False) -> Dict[str, Dict[str, str]]:
    """
    Carrega o mapeamento completo de verbas a partir:
    1. Do JSON de fixtures permanente (lojas/fixtures/de_para_verbas.json).
    2. Da planilha oficial apontada em PLANILHA_VERBAS_PATH (gera o JSON permanentemente).
    3. Do banco de dados SQLite (lojas.models.Verba) caso a planilha não esteja acessível.
    4. Do arquivo .example caso nenhuma das opções anteriores esteja disponível.

    Retorna um dicionário indexado tanto pelo código formatado com 3 dígitos (ex.: '001')
    quanto pelo código original sem zeros à esquerda (ex.: '1').
    """
    global _MAPA_DE_PARA
    if not forcar_recarga and _MAPA_DE_PARA is not None:
        return _MAPA_DE_PARA

    mapa: Dict[str, Dict[str, str]] = {}

    # 1. Se o JSON permanente existir e não for forçada a recarga
    if not forcar_recarga and os.path.exists(FIXTURE_PATH):
        try:
            with open(FIXTURE_PATH, "r", encoding="utf-8") as f:
                lista = json.load(f)
                # Se tiver mais de 10 verbas, é uma base completa válida
                if isinstance(lista, list) and len(lista) > 10:
                    for item in lista:
                        cod = str(item.get("codigo", "")).strip()
                        if not cod:
                            continue
                        cod_pad = cod.zfill(3) if cod.isdigit() else cod
                        desc = str(item.get("descricao", "")).strip()
                        tipo = normalizar_tipo_verba(item.get("tipo", "Provento"))

                        dados = {
                            "codigo": cod_pad,
                            "descricao": desc,
                            "tipo": tipo
                        }
                        mapa[cod_pad] = dados
                        raw = cod_pad.lstrip("0")
                        if raw and raw not in mapa:
                            mapa[raw] = dados

                    _MAPA_DE_PARA = mapa
                    return _MAPA_DE_PARA
        except Exception as e:
            print(f"Aviso ao ler fixture {FIXTURE_PATH}: {e}")

    # 2. Tenta gerar/carregar a partir da planilha Excel configurada
    caminho_excel = obter_caminho_planilha()
    if caminho_excel and os.path.exists(caminho_excel):
        try:
            import openpyxl
            wb = openpyxl.load_workbook(caminho_excel, data_only=True)
            ws = wb["VERBAS"] if "VERBAS" in wb.sheetnames else wb.active
            rows = list(ws.iter_rows(values_only=True))

            if rows and len(rows) > 1:
                header = [str(c).strip().lower() if c is not None else "" for c in rows[0]]
                idx_cod = 6
                idx_desc = 7
                idx_tipo = 9

                for i, h in enumerate(header):
                    if h in ("protheus", "codigo", "código", "cod_verba"):
                        idx_cod = i
                    elif h in ("descricao", "descrição", "descriçao") and idx_desc == 7:
                        idx_desc = i
                    elif h == "tipo":
                        idx_tipo = i

                mapa_itens_salvar: Dict[str, Dict[str, str]] = {}

                for r in rows[1:]:
                    if len(r) <= max(idx_cod, idx_desc, idx_tipo):
                        continue
                    cod = r[idx_cod]
                    desc = r[idx_desc]
                    tipo = r[idx_tipo]

                    if cod is not None and str(cod).strip():
                        cod_str = str(cod).strip()
                        cod_pad = cod_str.zfill(3) if cod_str.isdigit() else cod_str
                        desc_str = str(desc).strip() if desc is not None else ""
                        tipo_norm = normalizar_tipo_verba(str(tipo) if tipo is not None else "Provento")

                        item = {
                            "codigo": cod_pad,
                            "descricao": desc_str,
                            "tipo": tipo_norm
                        }
                        mapa_itens_salvar[cod_pad] = item
                        mapa[cod_pad] = item
                        raw = cod_pad.lstrip("0")
                        if raw and raw not in mapa:
                            mapa[raw] = item

                if mapa_itens_salvar:
                    lista_para_salvar = list(mapa_itens_salvar.values())
                    os.makedirs(os.path.dirname(FIXTURE_PATH), exist_ok=True)
                    with open(FIXTURE_PATH, "w", encoding="utf-8") as f:
                        json.dump(lista_para_salvar, f, ensure_ascii=False, indent=2)

                    _MAPA_DE_PARA = mapa
                    return _MAPA_DE_PARA
        except Exception as e:
            print(f"Aviso ao processar planilha Excel ({caminho_excel}): {e}")

    # 3. Fallback para a base de dados de Verba já importada no SQLite
    try:
        from lojas.models import Verba
        qs = Verba.objects.all()
        if qs.exists():
            mapa_itens_salvar = {}
            for v in qs:
                cod_str = str(v.codigo_verba).strip()
                if not cod_str:
                    continue
                cod_pad = cod_str.zfill(3) if cod_str.isdigit() else cod_str
                desc_str = v.descricao.strip()
                tipo_norm = normalizar_tipo_verba(v.tipo_codigo)

                item = {
                    "codigo": cod_pad,
                    "descricao": desc_str,
                    "tipo": tipo_norm
                }
                mapa_itens_salvar[cod_pad] = item
                mapa[cod_pad] = item
                raw = cod_pad.lstrip("0")
                if raw and raw not in mapa:
                    mapa[raw] = item

            if mapa_itens_salvar:
                lista_para_salvar = list(mapa_itens_salvar.values())
                os.makedirs(os.path.dirname(FIXTURE_PATH), exist_ok=True)
                with open(FIXTURE_PATH, "w", encoding="utf-8") as f:
                    json.dump(lista_para_salvar, f, ensure_ascii=False, indent=2)

                _MAPA_DE_PARA = mapa
                return _MAPA_DE_PARA
    except Exception as e:
        print(f"Aviso ao carregar verbas do banco SQLite: {e}")

    # 4. Fallback final para o arquivo .example caso nada mais esteja disponível
    if os.path.exists(FIXTURE_EXAMPLE_PATH):
        try:
            with open(FIXTURE_EXAMPLE_PATH, "r", encoding="utf-8") as f:
                lista = json.load(f)
                for item in lista:
                    cod = str(item.get("codigo", "")).strip()
                    desc = str(item.get("descricao", "")).strip()
                    tipo = normalizar_tipo_verba(item.get("tipo", "Provento"))
                    dados = {
                        "codigo": cod,
                        "descricao": desc,
                        "tipo": tipo
                    }
                    mapa[cod] = dados
                    raw = cod.lstrip("0")
                    if raw and raw not in mapa:
                        mapa[raw] = dados
            _MAPA_DE_PARA = mapa
            return _MAPA_DE_PARA
        except Exception as e:
            print(f"Aviso ao ler {FIXTURE_EXAMPLE_PATH}: {e}")

    _MAPA_DE_PARA = mapa
    return _MAPA_DE_PARA


def normalizar_tipo_verba(tipo_str: str) -> str:
    """
    Normaliza a classificação contábil:
    - 'Provento', 'Proventos', 'provento' -> 'Provento'
    - 'Desconto', 'desconto' -> 'Desconto'
    - 'Base', 'BASE' -> 'Base'
    """
    t = (tipo_str or "").strip().lower()
    if "desconto" in t:
        return "Desconto"
    if "base" in t:
        return "Base"
    return "Provento"


def obter_info_verba(codigo: str, fallback_descricao: str = "", fallback_tipo: str = "Provento") -> Dict[str, str]:
    """
    Busca as informações oficiais da verba no de-para:
    - codigo (com padding)
    - descricao (Protheus Descrição)
    - tipo (Provento, Desconto, Base)
    """
    mapa = carregar_mapa_verbas()
    cod_str = str(codigo).strip()
    cod_pad = cod_str.zfill(3) if cod_str.isdigit() else cod_str

    info = mapa.get(cod_pad) or mapa.get(cod_str)
    if info:
        return {
            "codigo": cod_pad,
            "descricao": info.get("descricao") or fallback_descricao or f"Verba {cod_pad}",
            "tipo": normalizar_tipo_verba(info.get("tipo") or fallback_tipo)
        }

    return {
        "codigo": cod_pad,
        "descricao": fallback_descricao or f"Verba {cod_pad}",
        "tipo": normalizar_tipo_verba(fallback_tipo)
    }
